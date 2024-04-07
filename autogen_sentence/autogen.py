import pandas as pd
import openpyxl
import os

from openpyxl.workbook import Workbook
from transformers import AutoTokenizer, AutoModel
from randomgen import randomgenarticles

os.environ['CUDA_VISIBLE_DEVICES'] = '1'
MODEL_PATH = os.environ.get('MODEL_PATH', 'THUDM/chatglm3-6b-128k')
TOKENIZER_PATH = os.environ.get("TOKENIZER_PATH", MODEL_PATH)

tokenizer = AutoTokenizer.from_pretrained(TOKENIZER_PATH, trust_remote_code=True)
model = AutoModel.from_pretrained(MODEL_PATH, trust_remote_code=True, device_map="auto").eval()

#wb=openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\本地大模型\\爆款文案批量生成器\\爆款文案_2024.xlsx")
wb=openpyxl.load_workbook("/home/liuge/Desktop/本地大模型/爆款文案_2024.xlsx")
sheet=wb["爆款文案_2024"]
sheet_phrase=wb["吸睛话术"]
phrase=""
for row in range(2,sheet_phrase.max_row + 1):
    s = sheet_phrase.cell(row=row, column=2).value
    if not s: continue
    phrase = phrase + s + "\n"
print(phrase)

cnt = sheet['B2'].value
if not cnt:
    cnt = int(cnt)
print("将每条自动生成"+str(cnt)+"条")
offset = 8
for row in range(10,sheet.max_row + 1):
    cellS=sheet.cell(row=row,column=2)
    if cellS.value:
        print(cellS.value)
    query = ("我给你发一段话，你保持原来的意思不变，帮我去润色一下啊，润色的长度和原来的差不多。我希望你在润色的时候加入或参考我下面这些短语或短句，你需要把这些短语和短句非常丝滑的，套用到你润色的文本里面，因为这样的短语和短句在呃短视频文案里面就是爆款的短语短句，大家喜欢看到这样的词语或短句，我希望你能理解我的意思，我给你发大概100句这样的短语和短句，然后你根据这个文案的整体的意思，非常丝滑的融入一些你不能说所有的东西都融入进去，你去挑选一些适合这条文案的，保证整体的语句是特别呃丝滑整体。也特别通通畅，也通俗易懂。下面是你要参考的短语或短句；\n"
                + phrase + "下面是你要润色的文本：\n" + cellS.value)
    wb_new = Workbook()
    sheet_new = wb_new.active
    sheet_new.title = str(row-9)
    sheet_new.cell(row=2, column=1).value = sheet.cell(row=3,column=1).value
    sheet_new.cell(row=2, column=2).value = sheet.cell(row=3, column=2).value
    past_key_values, history = None, []
    sentences = cellS.value.split("。")
    for i in range(cnt):
        current_length = 0
        article = ""
        for j,s in enumerate(sentences):
            if not s: continue
            query = "我给你发一段话，你保持原来的意思不变，帮我去润色一下啊，润色的长度和原来的差不多。 下面是你要润色的文本：" + s
            print(s)
            current_length = 0
            sentence = ""
            for response, history, past_key_values in model.stream_chat(tokenizer, query, history=history, top_p=1,
                                                                        temperature=0.8,
                                                                        past_key_values=past_key_values,
                                                                        return_past_key_values=True):
                print(response[current_length:], end="", flush=True)
                current_length = len(response)
                sentence = response
            # if article and not article.endswith("。"):
            #     article = article + "。"
            #article = article + sentence
            #need do autoremove here
            print("remove before:"+sentence)
            if "润色" in sentence:
                if ":" in sentence:
                    v = sentence.split(":")
                    if "润色" in v[0]:
                        sentence = sentence[len(v[0])+1:]
                elif "：" in sentence:
                    v = sentence.split("：")
                    if "润色" in v[0]:
                        sentence = sentence[len(v[0]) + 1:]
            print("remove after:" + sentence)
            cell = sheet_new.cell(row=i+offset, column=j+1)
            cell.value = sentence
        #sheet_new['A'+str(i+1)] = article
        print()
        #wb_new.save('C:\\Users\\Administrator\\Desktop\\本地大模型\\爆款文案批量生成器\\爆款文案_2024_新文案_句级_'+str(row-9)+'.xlsx')
    path = '/home/liuge/Desktop/本地大模型/爆款文案_2024_新文案_句级_' + str(
            row - 9) + '.xlsx'
    wb_new.save(path)
    randomgenarticles(path)
