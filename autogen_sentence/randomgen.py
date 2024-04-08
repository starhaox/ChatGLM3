import sys
import random
import argparse
import pandas as pd
import openpyxl
import os
import time
from openpyxl.workbook import Workbook
def randomgenarticles(path, count):
    #wb=openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\本地大模型\\爆款文案批量生成器\\爆款文案_2024.xlsx")
    #path = "爆款文案_2024_新文案_句级_1.xlsx"
    start_time = time.time()
    wb=openpyxl.load_workbook(path)
    sheet=wb.active
    # wb_new = Workbook()
    # sheet_new = wb_new.active
    #print(sheet.max_column)
    rows = [[""] for i in range(sheet.max_column+1)]
    #print(len(rows))
    offset = 8
    for col in range(1, sheet.max_column + 1):
        for row in range(offset, sheet.max_row + 1):
            v = sheet.cell(row=row, column=col).value
            rows[col].append(v)
            #print(v)
        #print("next col")
    dest = count
    articles = []
    mx_c = sheet.max_column
    mx_r = sheet.max_row-offset+1
    i = 0
    s = set()
    while i < dest:
        article = ""
        key = ""
        for j in range(1, mx_c+1):
            idx = random.randint(1, mx_r)
            key=key+str(idx)+" "
            article = article+rows[j][idx]
        if key in s:
            print("发现重复，重新生成")
            continue
        s.add(key)
        print(str(i+1)+":"+article)
        i = i+1
        articles.append(article)

    # for i in range(dest):
    #     article = ""
    #     for j in range(1, mx_c+1):
    #         idx = random.randint(1, mx_r)
    #         article = article+rows[j][idx]
    #     print(str(i+1)+":"+article)
    #     articles.append(article)
    wb_new = Workbook()
    sheet_new = wb_new.active
    for i in range(len(articles)):
        cell = sheet_new.cell(row=i + 1, column=1)
        cell.value = articles[i]
    wb_new.save(path[:-5] +'_random.xlsx')
    end_time = time.time()
    print("共生成" + str(len(s)) + "条, 耗时"+str(end_time-start_time)+"秒")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--path", type=str, default="/home/liuge/Desktop/本地大模型/句级文案生成/爆款文案_2024_新文案_句级_1.xlsx")
    parser.add_argument("--count", type=int, default=10000)
    args = parser.parse_args()
    randomgenarticles(args.path, args.count)
