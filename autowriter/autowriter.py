import pandas as pd
import openpyxl
import os

from openpyxl.workbook import Workbook
from transformers import AutoTokenizer, AutoModel

os.environ['CUDA_VISIBLE_DEVICES'] = '1'
MODEL_PATH = os.environ.get('MODEL_PATH', 'THUDM/chatglm3-6b-128k')
TOKENIZER_PATH = os.environ.get("TOKENIZER_PATH", MODEL_PATH)
print("=============模型加载============")
tokenizer = AutoTokenizer.from_pretrained(TOKENIZER_PATH, trust_remote_code=True)
model = AutoModel.from_pretrained(MODEL_PATH, trust_remote_code=True, device_map="auto").eval()

#wb=openpyxl.load_workbook("C:\\Users\\Administrator\\Desktop\\本地大模型\\爆款文案批量生成器\\爆款文案_2024.xlsx")
#wb=openpyxl.load_workbook("/home/liuge/Desktop/本地大模型/句级文案生成/爆款文案_2024.xlsx")
wb=openpyxl.load_workbook("/home/liuge/Desktop/本地大模型/文案仿写器/爆款文案_2024.xlsx")
sheet=wb["爆款文案_2024"]
sheet_phrase=wb["吸睛话术"]
sheet_templete=wb["爆款文案模版"]
phrase=""
for row in range(2,sheet_phrase.max_row + 1):
    s = sheet_phrase.cell(row=row, column=2).value
    if not s: continue
    phrase = phrase + s + "\n"
print(phrase)

cnt = sheet['B2'].value
if not cnt:
    cnt = int(cnt)

templetes = []
for row in range(7, sheet_templete.max_row + 1):
    tmp = sheet_templete.cell(row=row, column=3).value
    templetes.append(tmp)

print("=============将每条卖点自动按模版列表生成对应文案============")
offset = 8
for row in range(10,sheet.max_row + 1):
    cellS=sheet.cell(row=row,column=2)
    if cellS.value:
        print("原始参考文案")
        print(cellS.value)
    ori = cellS.value
    num = sheet.cell(row=row,column=3).value
    wb_new = Workbook()
    sheet_new = wb_new.active
    sheet_new.title = str(row - 9)

    for i, temp in enumerate(templetes):
        if not temp:continue
        for j in range(cnt):
            idx = j + 1+i*cnt
            print("=============第" + str(idx) + "条============")
            query = (ori + "\n请套用以下文案:\""+temp+"\"生成"+str(num)+"字左右的文案")
            print(query)
            past_key_values, history = None, []
            current_length = 0
            article = ""
            current_length = 0
            sentence = ""
            for response, history, past_key_values in model.stream_chat(tokenizer, query, history=history, top_p=1,
                                                                        temperature=0.8,
                                                                        past_key_values=past_key_values,
                                                                        return_past_key_values=True):
                print(response[current_length:], end="", flush=True)
                current_length = len(response)
                sentence = response
            cell = sheet_new.cell(row=idx+offset, column=1)
            cell.value = sentence
            #sheet_new['A'+str(i+1)] = article
            print()
            #wb_new.save('C:\\Users\\Administrator\\Desktop\\本地大模型\\爆款文案批量生成器\\爆款文案_2024_新文案_句级_'+str(row-9)+'.xlsx')
    path = '/home/liuge/Desktop/本地大模型/文案仿写器/爆款文案_2024_新文案_仿写_' + str(
            row - 9) + '.xlsx'
    wb_new.save(path)
